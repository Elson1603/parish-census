"""Renders a ReportData into PDF, Excel (.xlsx), or CSV bytes."""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.services.chart_images import render_chart_png
from app.services.report_data import ReportData, ReportTable


def _build_table(table: ReportTable) -> Table:
    data = [table.headers] + table.rows
    pdf_table = Table(data, repeatRows=1)
    pdf_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return pdf_table


def render_pdf(report: ReportData) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=report.title,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(report.title, styles["Title"]),
        Paragraph(report.description, styles["Normal"]),
        Spacer(1, 0.2 * inch),
    ]

    if report.charts:
        for chart in report.charts:
            png_bytes = render_chart_png(chart)
            story.append(Image(io.BytesIO(png_bytes), width=4.5 * inch, height=2.6 * inch))
        story.append(Spacer(1, 0.2 * inch))

    if report.summary is not None and report.summary is not report.detail:
        story.append(Paragraph("Summary", styles["Heading2"]))
        story.append(_build_table(report.summary))
        story.append(Spacer(1, 0.3 * inch))

    if report.detail.rows:
        story.append(PageBreak() if report.charts or report.summary else Spacer(1, 0))
        story.append(Paragraph("Detail", styles["Heading2"]))
        story.append(_build_table(report.detail))
    elif report.summary is None:
        story.append(Paragraph("No data recorded yet.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def render_excel(report: ReportData) -> bytes:
    workbook = Workbook()

    def write_sheet(ws, table: ReportTable):
        ws.append(table.headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in table.rows:
            ws.append(row)
        for column_cells in ws.columns:
            length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    first_sheet = True
    if report.summary is not None and report.summary is not report.detail:
        ws = workbook.active
        ws.title = "Summary"
        write_sheet(ws, report.summary)
        first_sheet = False

    if report.detail.rows or first_sheet:
        ws = workbook.active if first_sheet else workbook.create_sheet("Detail")
        if first_sheet:
            ws.title = "Detail"
        write_sheet(ws, report.detail)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def render_csv(report: ReportData) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    table = report.detail if report.detail.rows else report.summary
    if table is not None:
        writer.writerow(table.headers)
        writer.writerows(table.rows)

    return buffer.getvalue().encode("utf-8-sig")
